from fastapi import APIRouter, Request, Form, Depends, HTTPException, status, Response, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from io import BytesIO
import pandas as pd
from typing import Optional
import os
import json
from app.core.sockets import active_connections
from app.db.mongodb import get_database
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

router = APIRouter()

# Setup templates directory
templates = Jinja2Templates(directory="app/templates")

# Mocked credentials from user request
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin12345"

admin_connections: list[WebSocket] = []

async def broadcast_admin_stats(mobile: str | None = None, is_online: bool | None = None):
    db = await get_database()
    active_now = len(active_connections)
    total_enrollments = await db["students"].count_documents({})
    completed_exams = await db["students"].count_documents({"status": "completed"})
    
    pending_exams = total_enrollments - completed_exams
    
    msg = {
        "type": "stats",
        "active_now": active_now,
        "total_enrollments": total_enrollments,
        "completed_exams": completed_exams,
        "pending_exams": pending_exams
    }
    
    # If a specific student's status changed, include full enriched data in the broadcast
    if mobile is not None:
        student = await db["students"].find_one({"mobile": mobile})
        if student:
            from app.utils.excel_utils import parse_exam_questions
            exam_data = parse_exam_questions("app/Questions/exam_questions.xlsx")
            
            # Enrich
            student["_id"] = str(student["_id"])
            total_correct = 0
            answered_count = len(student.get("answers", {}))
            for section in exam_data["sections"]:
                for q in section["questions"]:
                    if student.get("answers", {}).get(q["id"]) == q["correct"]:
                        total_correct += 1
            
            student["total_questions"] = exam_data["total_questions"]
            student["answered_count"] = answered_count
            student["correct_count"] = total_correct
            student["incorrect_count"] = answered_count - total_correct
            student["unattended_count"] = student["total_questions"] - answered_count
            student["total_score"] = total_correct
            student["is_online"] = is_online if is_online is not None else (mobile in active_connections)
            
            rem = student.get("remaining_seconds", 3600)
            mins, secs = divmod(int(rem), 60)
            student["time_remaining"] = f"{mins:02d}:{secs:02d}"
            
            msg["student_update"] = student
    
    for conn in admin_connections:
        try:
            await conn.send_json(msg)
        except:
            if conn in admin_connections:
                admin_connections.remove(conn)

@router.get("/students-list")
async def get_students_list(request: Request):
    if request.cookies.get("admin_session") != "authenticated":
        raise HTTPException(status_code=401)
    
    db = await get_database()
    students = await db["students"].find().to_list(length=1000)
    
    from app.utils.excel_utils import parse_exam_questions
    exam_data = parse_exam_questions("app/Questions/exam_questions.xlsx")
    
    for s in students:
        s["_id"] = str(s["_id"])
        total_correct = 0
        section_scores = {}
        for section in exam_data["sections"]:
            sec_correct = 0
            for q in section["questions"]:
                if s.get("answers", {}).get(q["id"]) == q["correct"]:
                    sec_correct += 1
                    total_correct += 1
            section_scores[section["name"]] = sec_correct
        
        answered_count = len(s.get("answers", {}))
        s["total_score"] = total_correct
        s["section_wise"] = section_scores
        s["violation_count"] = s.get("violation_count", 0)
        s["is_online"] = s["mobile"] in active_connections
        s["total_questions"] = exam_data["total_questions"]
        s["answered_count"] = answered_count
        s["correct_count"] = total_correct
        s["incorrect_count"] = answered_count - total_correct
        s["unattended_count"] = s["total_questions"] - answered_count
        
        rem = s.get("remaining_seconds", 3600)
        mins, secs = divmod(int(rem), 60)
        s["time_remaining"] = f"{mins:02d}:{secs:02d}"
        s["completed_at"] = s.get("completed_at").isoformat() if s.get("completed_at") else None
        
    return students

@router.websocket("/admin-ws")
async def admin_websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    admin_connections.append(websocket)
    try:
        # Send initial stats
        db = await get_database()
        active_now = len(active_connections)
        total_enrollments = await db["students"].count_documents({})
        completed_exams = await db["students"].count_documents({"status": "completed"})
        
        pending_exams = total_enrollments - completed_exams
        
        await websocket.send_json({
            "type": "stats", 
            "active_now": active_now,
            "total_enrollments": total_enrollments,
            "completed_exams": completed_exams,
            "pending_exams": pending_exams
        })
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        if websocket in admin_connections:
            admin_connections.remove(websocket)

@router.get("/", response_class=HTMLResponse)
async def admin_login_page(request: Request, error: Optional[str] = None):
    return templates.TemplateResponse("admin/login.html", {"request": request, "error": error})

@router.post("/login")
async def admin_login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
):
    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        response = RedirectResponse(url="/administrator/dashboard", status_code=status.HTTP_303_SEE_OTHER)
        # In a real app, we'd set a secure cookie or JWT here
        response.set_cookie(key="admin_session", value="authenticated")
        return response
    
    return templates.TemplateResponse(
        "admin/login.html", 
        {"request": request, "error": "Invalid username or password"}
    )

@router.get("/dashboard", response_class=HTMLResponse)
async def admin_dashboard(request: Request):
    # Basic check for our mock session
    if request.cookies.get("admin_session") != "authenticated":
        return RedirectResponse(url="/administrator/")
    
    db = await get_database()
    total_students = await db["students"].count_documents({})
    active_now = len(active_connections)
    completed_students = await db["students"].count_documents({"status": "completed"})
    pending_students = total_students - completed_students
    config = get_portal_config()
    
    return templates.TemplateResponse(
        "admin/dashboard.html", 
        {
            "request": request,
            "total_students": total_students,
            "active_now": active_now,
            "completed_students": completed_students,
            "pending_students": pending_students,
            "config": config
        }
    )

@router.get("/reset-password", response_class=HTMLResponse)
async def reset_password_page(request: Request, error: Optional[str] = None):
    # Basic check for our mock session
    if request.cookies.get("admin_session") != "authenticated":
        return RedirectResponse(url="/administrator/")
    
    return templates.TemplateResponse("admin/reset_password.html", {"request": request, "error": error})

@router.post("/reset-password")
async def reset_password(
    request: Request,
    old_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...)
):
    if request.cookies.get("admin_session") != "authenticated":
        return RedirectResponse(url="/administrator/")

    # Check old password
    if old_password != ADMIN_PASSWORD:
         return templates.TemplateResponse(
            "admin/reset_password.html", 
            {"request": request, "error": "Incorrect old password."}
        )

    if new_password != confirm_password:
        return templates.TemplateResponse(
            "admin/reset_password.html", 
            {"request": request, "error": "New passwords do not match."}
        )
    
    # In a real app, logic to update database here
    # For now, we just mock the success and redirect
    return templates.TemplateResponse(
        "admin/reset_password.html", 
        {"request": request, "success_msg": "Password updated successfully!"}
    )

@router.get("/logout")
async def logout():
    response = RedirectResponse(url="/administrator/")
    response.delete_cookie("admin_session")
    return response

CONFIG_PATH = "app/core/portal_config.json"

def get_portal_config():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r") as f:
            return json.load(f)
    return {}

def save_portal_config(config):
    with open(CONFIG_PATH, "w") as f:
        json.dump(config, f, indent=4)

@router.get("/manage-exams")
async def manage_exams_page(request: Request):
    if request.cookies.get("admin_session") != "authenticated":
        return RedirectResponse(url="/administrator/")
    
    config = get_portal_config()
    # Prepare instructions as a string for textarea
    instructions_str = "\n".join(config.get("instructions", []))
    
    return templates.TemplateResponse("admin/manage_exams.html", {
        "request": request,
        "config": config,
        "instructions_str": instructions_str
    })

@router.post("/manage-exams")
async def save_exam_config(
    request: Request,
    exam_title: str = Form(...),
    motivation_message: str = Form(...),
    instructions: str = Form(...),
    total_time: int = Form(...),
    start_time: str = Form(...),
    inspect_allow: str = Form(None),
    full_screen_mode: str = Form(None),
    exit_attempts: int = Form(3),
    real_time_backup: str = Form(None),
    allow_copy: str = Form(None),
    show_results: str = Form(None),
    admin_notification_duration: int = Form(5),
    countdown_before_exam: int = Form(30),
    show_animations: str = Form(None)
):
    if request.cookies.get("admin_session") != "authenticated":
        return RedirectResponse(url="/administrator/")
    
    config = get_portal_config()
    config.update({
        "exam_name": exam_title,
        "motivation_message": motivation_message,
        "instructions": [i.strip() for i in instructions.split("\n") if i.strip()],
        "total_time_minutes": total_time,
        "start_time": start_time,
        "inspect_allow": inspect_allow == "on",
        "enforce_fullscreen": full_screen_mode == "on",
        "exit_attempts_threshold": exit_attempts,
        "real_time_backup": real_time_backup == "on",
        "allow_copy": allow_copy == "on",
        "show_results": show_results == "on" if show_results else (show_results == "on" if request.method == "POST" else config.get('show_results', True)),
        "admin_notification_duration": admin_notification_duration,
        "countdown_before_exam": countdown_before_exam,
        "show_login_animations": show_animations == "on"
    })
    
    save_portal_config(config)
    
    return templates.TemplateResponse("admin/manage_exams.html", {
        "request": request,
        "config": config,
        "instructions_str": instructions,
        "success_msg": "Configuration saved successfully!"
    })

@router.post("/toggle-animations")
async def toggle_animations(request: Request):
    if request.cookies.get("admin_session") != "authenticated":
        raise HTTPException(status_code=401)
    data = await request.json()
    config = get_portal_config()
    config["show_login_animations"] = data.get("show_animations", True)
    save_portal_config(config)
    return {"status": "success"}

@router.post("/toggle-fullscreen")
async def toggle_fullscreen(request: Request):
    if request.cookies.get("admin_session") != "authenticated":
        raise HTTPException(status_code=401)
    data = await request.json()
    config = get_portal_config()
    config["enforce_fullscreen"] = data.get("enforce_fullscreen", False)
    save_portal_config(config)
    return {"status": "success"}

@router.get("/students")
async def get_students_page(request: Request):
    if request.cookies.get("admin_session") != "authenticated":
        return RedirectResponse(url="/administrator/")
    
    config = get_portal_config()
    db = await get_database()
    active_now = await db["students"].count_documents({"status": "active"})
    total_students = await db["students"].count_documents({})
    completed_students = await db["students"].count_documents({"status": "completed"})
    
    return templates.TemplateResponse("admin/students.html", {
        "request": request,
        "config": config,
        "active_now": active_now,
        "total_students": total_students,
        "completed_students": completed_students
    })

@router.get("/results")
async def get_results_page(request: Request):
    if request.cookies.get("admin_session") != "authenticated":
        return RedirectResponse(url="/administrator/")
    
    config = get_portal_config()
    db = await get_database()
    active_now = await db["students"].count_documents({"status": "active"})
    total_students = await db["students"].count_documents({})
    completed_students = await db["students"].count_documents({"status": "completed"})
    
    return templates.TemplateResponse("admin/results.html", {
        "request": request,
        "config": config,
        "active_now": active_now,
        "total_students": total_students,
        "completed_students": completed_students
    })



@router.post("/send-message")
async def send_student_message(request: Request):
    if request.cookies.get("admin_session") != "authenticated":
        raise HTTPException(status_code=401)
    
    data = await request.json()
    mobile = data.get("mobile")
    message = data.get("message")
    
    if mobile in active_connections:
        try:
            await active_connections[mobile].send_json({
                "type": "admin_message",
                "message": message
            })
            return {"status": "success"}
        except:
            return {"status": "error", "message": "Failed to send message"}
    return {"status": "error", "message": "Student offline"}

@router.post("/broadcast-message")
async def broadcast_message(request: Request):
    if request.cookies.get("admin_session") != "authenticated":
        raise HTTPException(status_code=401)
    
    data = await request.json()
    message = data.get("message")
    
    count = 0
    for mobile, ws in active_connections.items():
        try:
            await ws.send_json({
                "type": "admin_message",
                "message": message
            })
            count += 1
        except:
            pass
    return {"status": "success", "count": count}

@router.get("/student-backup/{mobile}")
async def get_student_backup(request: Request, mobile: str):
    if request.cookies.get("admin_session") != "authenticated":
        raise HTTPException(status_code=401)
    
    db = await get_database()
    student = await db["students"].find_one({"mobile": mobile})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
        
    from app.utils.excel_utils import parse_exam_questions
    exam_data = parse_exam_questions("app/Questions/exam_questions.xlsx")
    
    backup_data = []
    for section in exam_data["sections"]:
        sec_data = {"name": section["name"], "questions": []}
        for q in section["questions"]:
            ans = student.get("answers", {}).get(q["id"])
            sec_data["questions"].append({
                "text": q["text"],
                "answered": ans is not None,
                "student_answer": ans,
                "correct_answer": q["correct"],
                "is_correct": ans == q["correct"] if ans else False
            })
        backup_data.append(sec_data)
        
    return {
        "name": student["name"],
        "sections": backup_data,
        "status": student.get("status"),
        "violation_count": student.get("violation_count", 0),
        "violations_list": student.get("violations", [])
    }

@router.post("/log-violation")
async def log_violation(request: Request):
    # This endpoint is called from the student's exam page
    data = await request.json()
    mobile = request.cookies.get("student_mobile")
    if not mobile:
        return {"status": "error", "message": "No session"}
        
    db = await get_database()
    await db["students"].update_one(
        {"mobile": mobile},
        {
            "$inc": {"violation_count": 1},
            "$push": {
                "violations": {
                    "timestamp": datetime.now(),
                    "ip": request.client.host,
                    "type": data.get("type", "unknown"),
                    "reason": data.get("reason", "Suspicious activity detected")
                }
            }
        }
    )
    return {"status": "success"}

@router.delete("/delete-student/{mobile}")
async def delete_student(request: Request, mobile: str):
    if request.cookies.get("admin_session") != "authenticated":
        raise HTTPException(status_code=401)
    
    db = await get_database()
    result = await db["students"].delete_one({"mobile": mobile})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Student not found")
        
    return {"status": "success"}

@router.delete("/delete-all-students")
async def delete_all_students(request: Request):
    if request.cookies.get("admin_session") != "authenticated":
        raise HTTPException(status_code=401)
    
    db = await get_database()
    result = await db["students"].delete_many({})
    
    return {"status": "success", "deleted_count": result.deleted_count}

@router.get("/export-results")
async def export_results_excel(request: Request):
    if request.cookies.get("admin_session") != "authenticated":
        raise HTTPException(status_code=401)
    
    db = await get_database()
    students = await db["students"].find({"status": "completed"}).to_list(length=1000)
    
    config = get_portal_config()
    exam_title = config.get("exam_name", "Exam Results")
    
    from app.utils.excel_utils import parse_exam_questions
    exam_data = parse_exam_questions("app/Questions/exam_questions.xlsx")
    
    export_data = []
    for s in students:
        total_correct = 0
        section_scores = {}
        for section in exam_data["sections"]:
            sec_correct = 0
            for q in section["questions"]:
                if s.get("answers", {}).get(q["id"]) == q["correct"]:
                    sec_correct += 1
                    total_correct += 1
            section_scores[f"{section['name']} ({len(section['questions'])})"] = sec_correct
        
        # Calculate Accuracy
        total_q = exam_data.get("total_questions", 30)
        accuracy = f"{(total_correct / total_q * 100):.1f}%" if total_q > 0 else "0%"
        
        row = {
            "Student Name": s.get("name", "--") or "--",
            "Parent Name": s.get("parent_name", "--") or "--",
            "Mobile/ID": s.get("mobile", "--") or "--",
            "Stream": s.get("stream", "--") or "--",
            "Address": s.get("address", "--") or "--",
            "Overall Score": f"{total_correct}/{total_q}",
            "Accuracy": accuracy,
            "Completion Date/Time": s.get("completed_at").strftime("%Y-%m-%d %H:%M:%S") if s.get("completed_at") else "--"
        }
        
        # Add section scores
        for sec_name, score in section_scores.items():
            row[sec_name] = score
            
        export_data.append(row)
    
    if not export_data:
        # Create an empty row if no results
        export_data = [{
            "Student Name": "--",
            "Mobile/ID": "--",
            "Overall Score": "--",
            "Accuracy": "--",
            "Completion Date/Time": "--"
        }]

    df = pd.DataFrame(export_data)
    
    # Reorder columns: Name, Parent, Mobile, Stream, Address, Score, Accuracy, [Sections], Date
    cols = ["Student Name", "Parent Name", "Mobile/ID", "Stream", "Address", "Overall Score", "Accuracy"]
    section_cols = [c for c in df.columns if "(" in c and ")" in c and c not in cols]
    cols.extend(section_cols)
    cols.append("Completion Date/Time")
    
    df = df[cols]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write title in first row
        # We'll use a trick: write the title to cell A1, and start the dataframe from row 4
        df.to_excel(writer, index=False, sheet_name='Results', startrow=3)
        
        workbook = writer.book
        worksheet = writer.sheets['Results']
        
        # Add Exam Title
        from openpyxl.styles import Font, PatternFill, Alignment
        worksheet['A1'] = str(exam_title).upper()
        worksheet['A1'].font = Font(size=22, bold=True, color="1E293B")
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(df.columns))
        
        # Style headers
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[4]: # Row 4 contains headers (startrow=3 means dataframe starts at row 4)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        # Set column widths (auto-fit)
        for col_idx, column in enumerate(df.columns, 1):
            max_length = 0
            column_letter = worksheet.cell(row=4, column=col_idx).column_letter
            
            # Check header length
            max_length = max(len(str(column)), max_length)
            
            # Check cell content lengths
            for cell in worksheet[column_letter]:
                if cell.row < 4: continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 4)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Add filters
        worksheet.auto_filter.ref = f"A4:{worksheet.cell(row=4, column=len(df.columns)).column_letter}{len(df) + 4}"

    output.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="exam_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    }
    
    return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@router.get("/export-pdf")
async def export_results_pdf(request: Request):
    try:
        if request.cookies.get("admin_session") != "authenticated":
            raise HTTPException(status_code=401)
        
        db = await get_database()
        students = await db["students"].find({"status": "completed"}).to_list(length=1000)
        
        config = get_portal_config()
        # Forces the user's requested title
        exam_title = "CA CS ENTRANCE EXAM RESULTS"
        
        from app.utils.excel_utils import parse_exam_questions
        exam_data = parse_exam_questions("app/Questions/exam_questions.xlsx")
        
        sections = exam_data.get("sections", [])
        total_possible = exam_data.get("total_questions", 100)
        
        styles = getSampleStyleSheet()
        
        # Header row for table
        table_headers = ["NAME", "PARENT NAME", "MOBILE NO"]
        for sec in sections:
            table_headers.append(f"{sec['name'].upper()} ({sec['count']})")
        table_headers.append(f"TOTAL ({total_possible})")
        
        # Use Paragraphs for headers to enable wrapping if needed
        p_headers = [Paragraph(f"<b>{h}</b>", styles['Normal']) for h in table_headers]
        table_data = [p_headers]
        
        for s in students:
            total_correct = 0
            section_scores = []
            for section in sections:
                sec_correct = 0
                for q in section["questions"]:
                    if s.get("answers", {}).get(q["id"]) == q["correct"]:
                        sec_correct += 1
                        total_correct += 1
                section_scores.append(str(sec_correct))
            
            row = [
                Paragraph(s.get("name", "--") or "--", styles['Normal']),
                Paragraph(s.get("parent_name", "--") or "--", styles['Normal']),
                Paragraph(s.get("mobile", "--") or "--", styles['Normal']),
                *section_scores,
                str(total_correct)
            ]
            table_data.append(row)

        if len(table_data) == 1:
            table_data.append(["No results found"] + [""] * (len(table_headers) - 1))

        output = BytesIO()
        # Use landscape A4 for wider tables
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=30)
        elements = []
        
        # Custom Page Styles
        # Custom Page Styles for a Premium Look
        college_title_style = ParagraphStyle(
            'CollegeTitle',
            parent=styles['Normal'],
            fontSize=22,
            textColor=colors.HexColor("#800000"), # Professional Dark Brown
            fontName='Helvetica-Bold',
            alignment=TA_LEFT,
            leading=26, # Reduced from 35
            spaceAfter=6
        )
        
        college_subtitle_style = ParagraphStyle(
            'CollegeSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.black,
            fontName='Helvetica-Bold', # Match the bolder look in screenshot
            alignment=TA_LEFT,
            leading=16, # Reduced from 20
            spaceBefore=12 # Clear space after the bold line
        )
        
        accreditation_style = ParagraphStyle(
            'Accreditation',
            parent=styles['Normal'],
            fontSize=9.5,
            textColor=colors.HexColor("#0056b3"), # Professional blue
            fontName='Helvetica',
            alignment=TA_LEFT,
            leading=16, # Increased from 12
            spaceBefore=4
        )
        
        exam_title_style = ParagraphStyle(
            'ExamTitle',
            parent=styles['Normal'],
            fontSize=16,
            textColor=colors.HexColor("#1E293B"),
            alignment=TA_CENTER,
            fontName='Helvetica-Bold',
            spaceBefore=20,
            spaceAfter=8,
            leading=20
        )
        
        academic_year_style = ParagraphStyle(
            'AcademicYear',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor("#64748B"),
            alignment=TA_CENTER,
            fontName='Helvetica-Bold',
            spaceAfter=25
        )

        # 1. College Header with Logo
        # 1. College Header with Logo
        logo_path = "app/static/logo/clglogo.png"
        logo_img = "LOGO"
        if os.path.exists(logo_path):
            try:
                logo_img = Image(logo_path, width=1.15*inch, height=1.15*inch)
            except:
                pass

        # Text part of the header
        text_content = [
            [Paragraph("Dr. B. B. Hegde First Grade College, Kundapura", college_title_style)],
            [Paragraph("A Unit of Coondapur Education Society(R)", college_subtitle_style)],
            [Paragraph("Accredited by NAAC with B++ Grade, Affiliated to Mangalore University, Karnataka", accreditation_style)]
        ]
        text_table = Table(text_content, colWidths=[8*inch])
        text_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('TOPPADDING', (0,0), (-1,0), 0),    # No top padding for first row
            ('TOPPADDING', (0,1), (-1,1), 10),   # 10pt gap below the bold line for the subheading (Balanced)
            ('TOPPADDING', (0,2), (-1,2), 2),    # Tighter spacing for NAAC line
            ('BOTTOMPADDING', (0,0), (-1,0), 8), # 8pt gap above the bold line (Balanced)
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('LINEBELOW', (0,0), (0,0), 3.5, colors.HexColor("#1e293b")), # Very bold dark line
        ]))

        header_table = Table([[logo_img, text_table]], colWidths=[1.4*inch, 8.5*inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (1,0), (1,0), 15), # Added space between logo and text
        ]))
        
        elements.append(header_table)
        elements.append(HRFlowable(width="100%", thickness=2, color=colors.black, spaceBefore=4, spaceAfter=2))
        
        # 2. Main Title
        elements.append(Paragraph(exam_title, exam_title_style))
        
        # Academic Year subtitle
        current_year = datetime.now().year
        ay_text = f"ACADEMIC YEAR {current_year}-{str(current_year+1)[2:]}"
        elements.append(Paragraph(f"<b>{ay_text}</b>", academic_year_style))

        # 3. Results Table
        num_cols = len(table_headers)
        
        # Balanced widths for single-line content
        name_w = 1.6*inch
        parent_w = 1.5*inch
        mobile_w = 1.2*inch # Wider to fit 10-digit numbers in one line
        total_w = 0.9*inch # Wider to fit "TOTAL (100)" in one line
        
        # Give remaining space to section columns
        remaining_w = (10.5*inch) - (name_w + parent_w + mobile_w + total_w)
        num_sections = len(sections) if sections else 1
        sec_w = remaining_w / num_sections
        
        col_widths = [name_w, parent_w, mobile_w] + [sec_w] * (num_cols - 4) + [total_w]
        
        # Table Styling
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=8, # Smaller to ensure single-line headers
            textColor=colors.whitesmoke,
            fontName='Helvetica-Bold',
            alignment=TA_CENTER,
            leading=10,
            allowWidows=0,
            allowOrphans=0
        )
        
        content_style = ParagraphStyle(
            'ContentStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor("#334155"), # Navy Slate for differentiation
            fontName='Helvetica',
            alignment=TA_LEFT,
            leading=12
        )
        
        # Wrap all content in Paragraphs for better alignment and wrapping
        styled_table_data = []
        for i, row in enumerate(table_data):
            if i == 0:
                # Header row
                styled_table_data.append([Paragraph(f"<b>{h}</b>", header_style) if isinstance(h, str) else h for h in table_headers])
            else:
                # Content rows
                new_row = []
                for cell in row:
                    if isinstance(cell, Paragraph):
                        new_row.append(cell)
                    else:
                        new_row.append(Paragraph(str(cell), content_style))
                styled_table_data.append(new_row)

        t = Table(styled_table_data, colWidths=col_widths, repeatRows=1)
        
        # Subtle zebra striping and sharp borders
        table_style_list = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0F172A")), # Dark header
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.7, colors.HexColor("#94A3B8")), # Sharp borders
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]), # Balanced zebra differentiation
        ]
        
        # Left align Name, Parent, Mobile columns (0, 1, 2)
        for row_idx in range(1, len(styled_table_data)):
            table_style_list.append(('ALIGN', (0, row_idx), (2, row_idx), 'LEFT'))

        t.setStyle(TableStyle(table_style_list))
        
        elements.append(t)
        
        # Build PDF and return
        doc.build(elements)
        output.seek(0)
        
        resp_headers = {
            'Content-Disposition': f'attachment; filename="exam_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        }
        return Response(content=output.getvalue(), media_type="application/pdf", headers=resp_headers)
        
    except Exception as e:
        import traceback
        error_msg = f"PDF EXPORT ERROR: {str(e)}\n\n{traceback.format_exc()}"
        print(error_msg)
        return Response(content=error_msg, status_code=500, media_type="text/plain")
